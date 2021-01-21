import os
#import utm
from openpyxl import load_workbook

workbook = load_workbook(filename='provascript2.xlsx')
nomiFogli = workbook.sheetnames
numSh = len(nomiFogli)
print("\tActive sheet: " + workbook.active.title)
print("\tSheets number: " + str(numSh))
foglio = workbook.active
coords = []
a=0
csv_file = open("pluvosta260.csv", "w")
stringa = ""
for x in range(8,foglio.max_row+1):
    for y in range(3,261):
        a+=1
        if(str(foglio.cell(x,y).value) == "None" or str(foglio.cell(x,y).value) == "-9999"):
            vals = 0
        else:
            vals = foglio.cell(x,y).value
        #print("ID: "+str(foglio.cell(1,y).value)+" mese: "+str(foglio.cell(x,1).value)+" anno: "+str(foglio.cell(x,2).value)+" | "+str(vals))
        #for a in range(0,foglio.max_row-7):
        stringa += str(a)+";"+str(foglio.cell(x,1).value)+";"+str(foglio.cell(x,2).value)+";"+str(vals)+";"+str(foglio.cell(1,y).value)+"\n"
csv_file.write(stringa)
csv_file.close()
    #print(str(foglio.cell(x,2).value) + "-" + str(foglio.cell(x,3).value))
    #if(foglio.cell(x,2).value != None or foglio.cell(x,3).value != None):
        #print(utm.to_latlon(foglio.cell(x,2).value, foglio.cell(x,3).value, 33, 'U'))
    #else:
        #print("(None,None)")