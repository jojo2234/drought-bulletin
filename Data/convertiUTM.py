import os
import utm
from openpyxl import load_workbook

workbook = load_workbook(filename='convertimi.xlsx')
nomiFogli = workbook.sheetnames
numSh = len(nomiFogli)
print("\tActive sheet: " + workbook.active.title)
print("\tSheets number: " + str(numSh))
foglio = workbook.active
coords = []
for x in range(2,foglio.max_row+1):
    #print(str(foglio.cell(x,2).value) + "-" + str(foglio.cell(x,3).value))
    if(foglio.cell(x,2).value != None or foglio.cell(x,3).value != None):
        print(utm.to_latlon(foglio.cell(x,2).value, foglio.cell(x,3).value, 33, 'U'))
    else:
        print("(None,None)")