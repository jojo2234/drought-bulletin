#Passata una cartella a questo script lui cercherà al suo interno i file excel ed estrarrà le stazioni meteorologiche
#Le inserirà in un file csv chiamato stazioni.csv da usare per aggiornare il database
#Se la stazione è già presente nel file csv o in memoria (altro file stessa esecuzione) allora non la ignora altrimenti la aggiunge
#Se il file csv non è presente lo crea dove si trova il programma
#C:\Users\Alex\Documents\GitHub\drought-bulletin\Data\PluvioData qui è dove si trovano i miei dati pluviometrici dove sono riportate le stazioni
#Per avviare il programma scrivere su cmd >python EstraiStazioniDaCartella.py [DirectoryConIFile]
import os
import sys
import xlrd
import string
from openpyxl import load_workbook

def do_extractionfrom(directoryentry):
    if directoryentry.name.endswith('xlsx'):
        workbook = load_workbook(filename=directoryentry)
        nomiFogli = workbook.sheetnames
        numSh = len(nomiFogli)
        print("\tActive sheet: " + workbook.active.title)
        print("\tSheets number: " + str(numSh))
        for i in range(0,numSh):
            trovato=False
            workbook.active = i
            print("\tOperating on sheet: " + workbook.active.title)
            foglio = workbook.active
            for x in range(1, foglio.max_row):
                if(trovato==True):
                    break
                for y in range(1, foglio.max_column):
                    celVal=""
                    if(type(foglio.cell(x,y).value) == str):
                        celVal = foglio.cell(x,y).value.lower()
                    if(celVal =="stazione" or celVal=="station" or celVal=="stazioni" or celVal=="stations"):
                        trovato=True
                        nomiStat = []
                        xcoord = []
                        ycoord = []
                        distr = []
                        quota = []
                        if("X_" not in foglio.cell(x,y+1).value and "cod_" not in foglio.cell(x,y+1).value and foglio.cell(x,y+1).value.isnumeric() == False):
                            #La cella dopo celVal molto problabilmente contine i nomi delle stazioni (sono in riga forse)
                            origY = y
                            for i in range(y+1,foglio.max_column):
                                if(i is not None):
                                    nomiStat.append(foglio.cell(x,i).value)
                                y+=1
                            x+=1
                            y = origY
                            if("x_" in str(foglio.cell(x,y).value.lower())):
                                for i in range(y+1,foglio.max_column):
                                    if(i is not None):
                                        xcoord.append(foglio.cell(x,i).value)
                                    y+=1
                            x+=1
                            y = origY
                            if("y_" in str(foglio.cell(x,y).value.lower())):
                                for i in range(y+1,foglio.max_column):
                                    if(i is not None):
                                        ycoord.append(foglio.cell(x,i).value)
                                    y+=1
                            x+=1
                            y = origY
                            if("distr" in str(foglio.cell(x,y).value.lower())):
                                for i in range(y+1,foglio.max_column):
                                    if(i is not None):
                                        distr.append(foglio.cell(x,i).value)
                                    y+=1
                            x+=2
                            y = origY
                            if("dtm" in str(foglio.cell(x,y).value.lower())):
                                for i in range(y+1,foglio.max_column):
                                    if(i is not None):
                                        quota.append(foglio.cell(x,i).value)
                                    y+=1
                            y = origY
                            csv_file = open("stazioni_"+foglio.title+".csv", "w")
                            stringa = ""
                            for a in range(0,len(nomiStat)):
                                stringa += str(a+1)+";"+str(xcoord[a])+";"+str(ycoord[a])+";"+str(nomiStat[a])+";"+str(quota[a])+";"+str(distr[a])+"\n"
                            csv_file.write(stringa)
                            csv_file.close()
                            break
                            #Si suppone che il resto dei dati se i nomi sono in riga siano sulla colonna
                        else:
                            #La cella dopo celVal su y non contiene i nomi delle stazioni quindi procedere su x
                            origX = x
                            for i in range(x+1,foglio.max_row):
                                if(i is not None):
                                    nomiStat.append(foglio.cell(i,y).value)
                                x+=1
                            y+=1
                            x = origX
                            if("x_" in str(foglio.cell(x,y).value.lower())):
                                for i in range(x+1,foglio.max_row):
                                    if(i is not None):
                                        xcoord.append(foglio.cell(i,y).value)
                                    x+=1
                            y+=1
                            x = origX
                            if("y_" in str(foglio.cell(x,y).value.lower())):
                                for i in range(x+1,foglio.max_row):
                                    if(i is not None):
                                        ycoord.append(foglio.cell(i,y).value)
                                    x+=1
                            y+=1
                            x = origX
                            if("distr" in str(foglio.cell(x,y).value.lower())):
                                for i in range(x+1,foglio.max_row):
                                    if(i is not None):
                                        distr.append(foglio.cell(i,y).value)
                                    x+=1
                            y+=2
                            x = origX
                            if("dtm" in str(foglio.cell(x,y).value.lower())):
                                for i in range(x+1,foglio.max_row):
                                    if(i is not None):
                                        quota.append(foglio.cell(i,y).value)
                                    x+=1
                            x = origX
                            csv_file = open("stazioni_"+foglio.title+".csv", "w")
                            stringa = ""
                            for a in range(0,len(nomiStat)):
                                stringa += str(a+1)+";"+str(xcoord[a])+";"+str(ycoord[a])+";"+str(nomiStat[a])+";"+str(quota[a])+";"+str(distr[a])+"\n"
                            csv_file.write(stringa)
                            csv_file.close()
                            break
                            #Si suppone che il resto dei dati se i nomi sono in colonna siano sulla riga
                            #Notabene nelle liste ci sono dei duplicati da rimuoverli inserendo la lista in un dict
                    #print(foglio.cell(x,y).value)
            #for row in foglio.iter_rows(min_row=1,max_row=8,min_col=1,max_col=20):
            #    print(row.value) #Errato?
            #print("\n\t " + str(foglio["B3"].value))
        print("\n")
    elif directoryentry.name.endswith('xls'):
        book = xlrd.open_workbook(directoryentry)
        numsht = book.nsheets
        print("\tSheets number: " + str(numsht))
        for i in range(0,numsht):
            trovato = False
            sht = book.sheet_by_index(i)
            print("\tOperating on sheet: " + sht.name)
            for x in range(1,sht.nrows):
                if(trovato==True):
                    break
                for y in range(1,sht.ncols):
                    celVal=""
                    if(type(sht.cell(x,y).value) == str):
                        celVal = sht.cell(x,y).value.lower()
                    if(celVal =="stazione" or celVal=="station" or celVal=="stazioni" or celVal=="stations"):
                        trovato = True
                        nomiStat = []
                        xcoord = []
                        ycoord = []
                        distr = []
                        quota = []
                        if("X_" not in sht.cell(x,y+1).value and "cod_" not in sht.cell(x,y+1).value and sht.cell(x,y+1).value.isnumeric() == False):
                            #print(sht.cell_value(x,y))
                            origY = y
                            for i in range(y+1,sht.ncols):
                                if(i is not None):
                                    nomiStat.append(sht.cell(x,i).value)
                                y+=1
                            x+=1
                            y = origY
                            if("x_" in str(sht.cell(x,y).value.lower())):
                                for i in range(y+1,sht.ncols):
                                    if(i is not None):
                                        xcoord.append(sht.cell(x,i).value)
                                    y+=1
                            x+=1
                            y = origY
                            if("y_" in str(sht.cell(x,y).value.lower())):
                                for i in range(y+1,sht.ncols):
                                    if(i is not None):
                                        ycoord.append(sht.cell(x,i).value)
                                    y+=1
                            x+=1
                            y = origY
                            if("distr" in str(sht.cell(x,y).value.lower())):
                                for i in range(y+1,sht.ncols):
                                    if(i is not None):
                                        distr.append(sht.cell(x,i).value)
                                    y+=1
                            x+=2
                            y = origY
                            if("dtm" in str(sht.cell(x,y).value.lower())):
                                for i in range(y+1,sht.ncols):
                                    if(i is not None):
                                        quota.append(sht.cell(x,i).value)
                                    y+=1
                                y = origY
                            #Si suppone che il resto dei dati se i nomi sono in riga siano sulla colonna
                            csv_file = open("stazioni_"+sht.name+".csv", "w")
                            stringa = ""
                            for a in range(0,len(nomiStat)):
                                stringa = (a+1)+";"+xcoord[a]+";"+ycoord[a]+";"+nomiStat[a]+";"+quota[a]+";"+distr[a]+"\n"
                            csv_file.write(stringa)
                            csv_file.close()
                            break
                        else:
                            #La cella dopo celVal su y non contiene i nomi delle stazioni quindi procedere su x
                            origX = x
                            for i in range(x+1,sht.nrows):
                                if(i is not None):
                                    nomiStat.append(sht.cell(i,y).value)
                                x+=1
                            y+=1
                            x = origX
                            if("x_" in str(sht.cell(x,y).value.lower())):
                                for i in range(x+1,sht.nrows):
                                    if(i is not None):
                                        xcoord.append(sht.cell(i,y).value)
                                    x+=1
                            y+=1
                            x = origX
                            if("y_" in str(sht.cell(x,y).value.lower())):
                                for i in range(x+1,sht.nrows):
                                    if(i is not None):
                                        ycoord.append(sht.cell(i,y).value)
                                    x+=1
                            y+=1
                            x = origX
                            if("distr" in str(sht.cell(x,y).value.lower())):
                                for i in range(x+1,sht.nrows):
                                    if(i is not None):
                                        distr.append(sht.cell(i,y).value)
                                    x+=1
                            y+=2
                            x = origX
                            if("dtm" in str(sht.cell(x,y).value.lower())):
                                for i in range(x+1,sht.nrows):
                                    if(i is not None):
                                        quota.append(sht.cell(i,y).value)
                                    x+=1
                            x = origX
                            csv_file = open("stazioni_"+sht.name+".csv", "w")
                            stringa = ""
                            for a in range(0,len(nomiStat)):
                                stringa = (a+1)+";"+xcoord[a]+";"+ycoord[a]+";"+nomiStat[a]+";"+quota[a]+";"+distr[a]+"\n"
                            csv_file.write(stringa)
                            csv_file.close()
                            break
    else:
        print("\tFile not supported")

if(len(sys.argv) > 1):
    dire = sys.argv[1]
else:
    dire = input("\nDirectory where xlsx files are located: ")

assert os.path.exists(dire), "Directory doesn't exist!"
print("\nSelected directory is: " + dire)
print("\nProcessing files: ")
with os.scandir(dire) as entries:
    for entry in entries:
        print("\n --- " + entry.name)
        do_extractionfrom(entry)