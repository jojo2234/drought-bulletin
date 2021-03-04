#Passata una cartella a questo script lui cercherà al suo interno i file excel ed estrarrà le stazioni meteorologiche
#Le inserirà in un file csv chiamato stazioni.csv da usare per aggiornare il database
#Se la stazione è già presente nel file csv o in memoria (altro file stessa esecuzione) allora non la ignora altrimenti la aggiunge
#Se il file csv non è presente lo crea dove si trova il programma
#C:\Users\Alex\Documents\GitHub\drought-bulletin\Data\PluvioData qui è dove si trovano i miei dati pluviometrici dove sono riportate le stazioni
#Per avviare il programma scrivere su cmd >python EstraiStazioniDaCartella.py [DirectoryConIFile]
import os
import sys
import xlrd
import string
from openpyxl import load_workbook
import datetime, time

def do_extractionfrom(directoryentry):
    if directoryentry.name.endswith('xlsx'):
        workbook = load_workbook(filename=directoryentry)
        nomiFogli = workbook.sheetnames
        numSh = len(nomiFogli)
        print("\tActive sheet: " + workbook.active.title)
        print("\tSheets number: " + str(numSh) + "\n")
        foglio = workbook.active
        a=0
        csv_file = open(directoryentry.name + ".csv", "w")
        stringa = ""
        for x in range(8,foglio.max_row+1):
            for y in range(3,foglio.max_column+1):
                a+=1
                if(str(foglio.cell(x,y).value) == "None" or str(foglio.cell(x,y).value) == "-9999"):
                    vals = 0
                else:
                    vals = foglio.cell(x,y).value
                #stringa = idPreci;mese;anno;valorePerY;IdStazioneY
                stringa += str(a)+";"+str(foglio.cell(x,1).value)+";"+str(foglio.cell(x,2).value)+";"+str(vals)+";"+str(foglio.cell(1,y).value)+"\n"
        csv_file.write(stringa)
        csv_file.close()
    elif directoryentry.name.endswith('xls'):
        book = xlrd.open_workbook(directoryentry)
        numsht = book.nsheets
        print("\tSheets number: " + str(numsht))
        sht = book.sheet_by_index(0)
        print("\tOperating on sheet: " + sht.name)
        a = 0 #Id precipitazione giornaliera
        csv_file = open(directoryentry.name + ".csv", "w")
        stringa = ""
        if("Y" in str(sht.cell(9,0).value)):
            starX = 10
        else:
            starX = 8
        if(type(sht.cell(16,0).value)==int and type(sht.cell(16,1).value)==int): #Controlla se la data è anno,anno anno,valore pultroppo questa cosa varia a seconda del file delle volte la data è un stringa perchè hanno messo un apice davanti altre volte è un float altre volte mese anno e giorno sono separati e possono venire letti come float e come string per ora mai come int quindi questo metodo non funziona
            starY = 3
        else:
            starY = 1
        for x in range(starX,sht.nrows):
            idStaz = 9900000
            for y in range(starY,sht.ncols):
                idStaz += 1 #Perchè in un file manca l'ID
                a+=1
                if(str(sht.cell(x,y).value) == "None" or str(sht.cell(x,y).value) == "-9999" or str(sht.cell(x,y).value) == "" or str(sht.cell(x,y).value) == "-9999.0"):
                    vals = 0
                else:
                    vals = sht.cell(x,y).value
                    #Questo file ha anno mese e giorno separati in questordine
                    idSta = sht.cell(1,y).value
                    if(idSta == None or idSta==""):
                        idSta = 0
                    #idprecipitazione;anno-mese-giorno;valore;idStazione
                    data = sht.cell(x,0).value
                    if(starY == 3): #In realtà dovrebbe automaticamente individuare dove e come sono posizionati i dati starX=1 pluvio file 143
                        #print("Line: " + str(x)) #All'interno dell'if controlla il file 71
                        stringa += str(a)+";"+str(int(sht.cell(x,0).value))+"-"+str(int(sht.cell(x,1).value))+"-"+str(int(sht.cell(x,2).value))+";"+str(vals)+";"+str(int(idSta))+"\n"
                        print(stringa)
                    else:
                        if(type(data) == float):
                            seconds = (data - 25569) * 86400.0
                            data = datetime.datetime.utcfromtimestamp(seconds).strftime('%Y-%m-%d')
                        stringa += str(a)+";"+str(data)+";"+str(vals)+";"+str(idStaz)+"\n"
        csv_file.write(stringa)
        csv_file.close()
    else:
        print("\tFile not supported")

if(len(sys.argv) > 1):
    dire = sys.argv[1]
else:
    dire = input("\nDirectory where xlsx files are located: ")

assert os.path.exists(dire), "Directory doesn't exist!"
print("\nSelected directory is: " + dire)
print("\nProcessing files: ")
with os.scandir(dire) as entries:
    for entry in entries:
        print("\n --- " + entry.name)
        do_extractionfrom(entry)